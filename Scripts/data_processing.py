from openpyxl import load_workbook
from generate_analytical_results import generate_analytical_results
from generate_incoice import generate_invoice
from generate_job_cover import generate_job_cover
from generate_cover_letter import generate_cover_letter
from tally import increase_tally
from datetime import date

def process_file(file_path, countryDropdown, typeDropdown):
    # Load the Excel file using openpyxl
    wb = load_workbook(file_path)
    
    # Generate output files
    # Extract data
    header_info = extract_header(wb)
    client_job_info = extract_client_job(wb)
    sample_info = extract_samples(wb)
    additional_comments_info = extract_additional_comments(wb)
    handle_log_info = extract_handle_log(wb)
    
    coc_data = {
        'header_info': header_info,
        'client_job_info': client_job_info,
        'sample_info': sample_info,
        'additional_comments_info': additional_comments_info,
        'handle_log_info': handle_log_info
    }
    
    #generate_analytical_results(coc_data, countryDropdown)
    generate_invoice(coc_data, countryDropdown, typeDropdown)
    generate_job_cover(coc_data, countryDropdown, typeDropdown)
    generate_cover_letter(coc_data, countryDropdown, typeDropdown)
    
    increase_tally()

def extract_header(wb):
    header_info = {
        'us_address': None,
        'can_address': None,
        'phone': None,
        'email': None
    }

    for ws in wb.worksheets:
        if header_info['us_address'] is None:
            header_info['us_address'] = ws.cell(row=3, column=5).value
        
        if header_info['can_address'] is None:
            header_info['can_address'] = ws.cell(row=6, column=5).value
        
        if header_info['phone'] is None:
            header_info['phone'] = ws.cell(row=6, column=1).value
        
        if header_info['email'] is None:
            header_info['email'] = ws.cell(row=7, column=1).value

        # Exit loop if all data has been found
        if all(header_info.values()):
            break
    
    return header_info

def extract_client_job(wb):
    # Function to check if a field is empty
    def check_field(field_name, value):
        if value is None or str(value).strip() == "":
            raise ValueError(f"'{field_name}' is empty")
    
    reference_data = None

    for ws in wb.worksheets:
        if is_sheet_blank(ws):
            continue
        
        current_data = {
            'client': ws.cell(row=8, column=2).value,
            'address': ws.cell(row=9, column=2).value,
            'project': ws.cell(row=10, column=2).value,
            'contact': convert_contact(ws.cell(row=11, column=2).value),
            'phone': ws.cell(row=12, column=2).value,
            'email': ws.cell(row=13, column=2).value,
            '#_samples': ws.cell(row=10, column=9).value,
            'job_#': ws.cell(row=11, column=9).value,
            'analysis': extract_highlighted_value(ws, 12, 9, 10),
            'photograph': extract_highlighted_value(ws, 13, 9, 10)
        }

        # Check if any field is empty
        for field_name, value in current_data.items():
            check_field(field_name, value)

        if reference_data is None:
            reference_data = current_data
        else:
            # Compare the current sheet's data with the reference data
            for key, value in current_data.items():
                if reference_data[key] != value:
                    raise ValueError(f"Discrepancy found in '{key}' between worksheets: {value} != {reference_data[key]}")
    
    return reference_data

def extract_samples(wb):
    data = []
    
    acceptable_matrix_codes = {
        "A",
        "BA",
        "BAF",
        "AOC",
        "RCS",
        "B",
        "BF",
        "S",
        "SF",
        "OT"
    }
    
    # Loop through each sheet (page) in the workbook
    for ws in wb.worksheets:
        # Loop through rows 17 to 32 on each sheet
        for row in range(17, 33):
            sample_id = ws.cell(row=row, column=1).value
            sample_identification = ws.cell(row=row, column=2).value
            matrix = ws.cell(row=row, column=7).value
            flow_rate = ws.cell(row=row, column=8).value
            time = ws.cell(row=row, column=9).value
            vol_area = ws.cell(row=row, column=10).value

            # Skip the row if the first cell (column 1) is empty
            if sample_id is None:
                continue

            # Check if the matrix value is in the acceptable matrix codes
            if matrix not in acceptable_matrix_codes:
                raise ValueError(f"Invalid matrix code '{matrix}' found in row {row} on sheet '{ws.title}'.")

            # Store the sample data
            data.append({
                'sample#': sample_id,
                'sample_identification': sample_identification,
                'matrix': matrix,
                'flow_rate': flow_rate,
                'time': time,
                'vol_area': vol_area
            })

    return data

def extract_additional_comments(wb):
    data = ""

    # Loop through each worksheet
    for ws in wb.worksheets:
        row = 1

        # Find the row where "Additional Comments and Instructions" starts
        while ws.cell(row=row, column=1).value != "Additional Comments and Instructions":
            row += 1
            if row > 100:
                break  # Exit if the end of the sheet is reached without finding the section

        # Extract the comment next to the "Additional Comments and Instructions" title
        comment = str(ws.cell(row=row, column=6).value or "")
        data += comment
        row += 1

        # Continue to collect comments below the title
        while True:
            comments = ws.cell(row=row, column=1).value
            
            if comments == " Chain of Custody Form":
                break
            if comments is not None:
                data += str(comments)
            
            if row > 100:
                break
            row += 1
        
        # If data is found, return it immediately
        if data.strip():  # Check if data is not empty or whitespace
            return data
    
    return data  # If no comments are found in any sheets, return an empty string or whatever was collected

def extract_handle_log(wb):
    
    sampled_by = {'name': None, 'signature': None, 'date_time': None}
    relinquished_by = {'name': None, 'signature': None, 'date_time': None}
    reviewed_by = {'name': None, 'signature': None, 'date_time': None}

    for ws in wb.worksheets:
        
        row = 1
        
        while ws.cell(row=row, column=1).value != " Chain of Custody Form":
            row += 1
            if row > 100:
                break 
        
        row += 1
        if sampled_by['name'] is None and sampled_by['signature'] is None and sampled_by['date_time'] is None:
            sampled_by = {
                'name': ws.cell(row=row, column=3).value,
                'signature': ws.cell(row=row, column=6).value,
                'date_time': ws.cell(row=row, column=9).value
            }
        
        row += 1
        if relinquished_by['name'] is None and relinquished_by['signature'] is None and relinquished_by['date_time'] is None:
            relinquished_by = {
                'name': ws.cell(row=row, column=3).value,
                'signature': ws.cell(row=row, column=6).value,
                'date_time': ws.cell(row=row, column=9).value
            }
        
        row += 1
        if reviewed_by['name'] is None and reviewed_by['signature'] is None and reviewed_by['date_time'] is None:
            reviewed_by = {
                'name': ws.cell(row=row, column=3).value,
                'signature': ws.cell(row=row, column=6).value,
                'date_time': ws.cell(row=row, column=9).value
            }
        
        if all([sampled_by['name'], sampled_by['signature'], sampled_by['date_time'],
                relinquished_by['name'], relinquished_by['signature'], relinquished_by['date_time'],
                reviewed_by['name'], reviewed_by['signature'], reviewed_by['date_time']]):
            break
    
    if reviewed_by['date_time'] is None:
        today = date.today()
        formatted_date = today.strftime("%A, %B %d, %Y")
        reviewed_by['date_time'] = formatted_date
    
    data = {
        'sampled_by': sampled_by,
        'relinquished_by': relinquished_by,
        'reviewed_by': reviewed_by
    }
    
    return data

def extract_highlighted_value(worksheet, row, col1, col2):
    cell1 = worksheet.cell(row=row, column=col1)
    cell2 = worksheet.cell(row=row, column=col2)

    fill1 = cell1.fill.fgColor.rgb if cell1.fill.fgColor else None
    fill2 = cell2.fill.fgColor.rgb if cell2.fill.fgColor else None

    # Check if both cells are highlighted
    if fill1 and fill1 != "00000000" and fill2 and fill2 != "00000000":
        if row == 12:
            raise ValueError(f"Both rush and standard is highlighted.")
        elif row == 13:
            raise ValueError(f"Both Yes and No is highlighted.")
        else:
            raise ValueError(f"Both cells in row {row} at columns {col1} and {col2} are highlighted.")
    
    # Return the value of the first highlighted cell
    if fill1 and fill1 != "00000000":
        return cell1.value
    elif fill2 and fill2 != "00000000":
        return cell2.value
    else:
        return None

def is_sheet_blank(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                return False  # Sheet is not blank
    return True  # Sheet is blank

def convert_contact(name):
    # Strip any leading/trailing whitespace and commas
    name = name.strip().strip(',')

    # Define the name mappings
    names = {
        "Vas, Candice, Admin, Alex, Ryan, Mark": "Mr. Vas Kanellos",
        "Vas": "Mr. Vas Kanellos",
        "Mark": "Mr. Vas Kanellos",
        "Candice": "Ms. Candice Rogers",
        "Alex": "Mr. Alex Fisher",
        "Ryan": "Mr. Ryan Boddis",
        "Admin": "Admin"
    }
    
    if name in names:
        return names[name]
    else:
        raise ValueError(f"The name '{name}' is not recognizable as a contact.")

    
def printData(coc_data):
    
    print("\nHEADER INFO ------------------------------------\n")
    
    for key, value in coc_data['header_info'].items():
        print(f'{key}: {value}')
        
    print("\nCLIENT JOB INFO ------------------------------------\n")
        
    for key, value in coc_data['client_job_info'].items():
        print(f'{key}: {value}')
        
    print("\nSAMPLE INFO ------------------------------------\n")
        
    for x in coc_data['sample_info']:
        print(x)
        
    print("\nADDITIONAL COMMENTS INFO ------------------------------------\n")
        
    print(coc_data['additional_comments_info'])

    print("\nHANDLE LOG INFO ------------------------------------\n")

    for key, value in coc_data['handle_log_info'].items():
        print(f"{key}:")
        for sub_key, sub_value in value.items():
            print(f"  {sub_key}: {sub_value}")
