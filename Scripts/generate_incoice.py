from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from tally import read_tally
from datetime import datetime
import os
import sys
import platform

global header_info, client_job_info, sample_info, handle_log_info
header_info = None
client_job_info = None
sample_info = None
handle_log_info = None

global ws, country, type
ws = None  
country = None
type = None

global starting_row, end_row
starting_row = None
end_row = None

def get_asset_path(filename):
    if getattr(sys, 'frozen', False):
        # When running as a packaged executable
        base_path = sys._MEIPASS
    else:
        return f'../{filename}'
    
    return os.path.join(base_path, filename)

def get_before_comma(input_string):
    # Check if a comma exists in the string
    if ',' in input_string:
        # Split the string at the first comma and return the part before the comma
        return input_string.split(',', 1)[0].strip()
    else:
        # If no comma, return the entire string
        return input_string.strip()

def sides(row):
    global ws
    cell_left = ws[f'A{row}']
    cell_right = ws[f'K{row}']
    
    if cell_left.border and cell_left.border.top and cell_left.border.top.style:
        cell_left.border = Border(left=Side(style='double'), top=Side(style='double'))
    elif cell_left.border and cell_left.border.bottom and cell_left.border.bottom.style:
        cell_left.border = Border(left=Side(style='double'), bottom=Side(style='thin'))
    else:
        cell_left.border = Border(left=Side(style='double'))
    
    cell_right.border = Border(left=Side(style='double'))

def line_across(row):
    global ws
    ws.cell(row=row, column=1).border = Border(top=Side(style='double'), left=Side(style='double'))
    for col in range(2, 11):
        ws.cell(row=row, column=col).border = Border(top=Side(style='double'))
        
def add_styling():
    global starting_row, end_row
    
    line_across(starting_row)
    for row in range (starting_row, end_row):
        sides(row)
    
    ws.cell(row=starting_row, column=5).border = Border(left=Side(style='double'), top=Side(style='double'), bottom=Side(style='double'))
    ws.cell(row=starting_row, column=10).border = Border(right=Side(style='double'), top=Side(style='double'), bottom=Side(style='double'))
    for col in range(6, 10):
        ws.cell(row=starting_row, column=col).border = Border(bottom=Side(style='double'), top=Side(style='double'))

def setup_excel():
    global ws
    # Set row heights
    ws.row_dimensions[1].height = 35.25
    
    # Set column widths
    ws.column_dimensions['A'].width = 10.86 + 0.85 # 10.71 (0.65) 10.89 (0.85)
    ws.column_dimensions['B'].width = 8.43 + 0.75 # 8.44 (0.75)
    ws.column_dimensions['C'].width = 8.43 + 0.75
    ws.column_dimensions['D'].width = 8.43 + 0.75
    ws.column_dimensions['E'].width = 9.86 + 0.85 # 9.78 (0.65) 9.89 (0.85)
    ws.column_dimensions['F'].width = 8.86 + 0.85 # 8.78 (0.65) 8.89 (0.85)
    ws.column_dimensions['G'].width = 2.71 + 0.85 # 2.67 (0.75) 2.78 (0.85)
    ws.column_dimensions['H'].width = 4.71 + 0.85 # 4.67 (0.75) 4.78 (0.85)
    ws.column_dimensions['I'].width = 10.86 + 0.85 # 10.78 (0.75) 10.89 (0.85)
    ws.column_dimensions['J'].width = 11.86 + 0.85 # 11.78 (0.75) 11.89 (0.85)

def invoice_number():
    global country
    today = datetime.now()
    year = str(today.year)[-2:]
    month = str(today.month).zfill(2)
    day = str(today.day).zfill(2)
    tally = read_tally()

    country_code = 'C' if country == 'Canada' else 'F'
    invoice_number = f"{country_code}{year}{month}{day}{tally}"

    return invoice_number

def project_number():
    global country, type
    country_code = 'C' if country == 'Canada' else 'F'
    tally = read_tally()
    now = datetime.now()
    current_year = str(now.year)[-2:]
    current_month = str(now.month).zfill(2)
    return f"{country_code}{type}{current_year}{current_month}{tally}"

def content_title():
    global type, client_job_info
    if type == 'MA':
        type_name = 'Microbiological Analysis'
    else: 
        type_name = 'Particle Sample'
        
    project = str(client_job_info['project'])    
    
    first = f"{type_name} for {project}"
    second = f"{client_job_info['client']} Project #{client_job_info['job_#']}"
    
    titles = {
        'first': first,
        'second': second
    }
    
    return titles
        
def split_text_based_on_column_width(start_cell, text, max_char_per_line):
    global ws
    number_of_rows_used = 0
    start_row = int(start_cell[1:])
    start_col = start_cell[0]
    words = text.split(' ')
    lines = []
    current_line = ""

    for word in words:
        if len(current_line) + len(word) + 1 > max_char_per_line:
            lines.append(current_line)
            current_line = word
        else:
            current_line += ' ' + word if current_line else word

    if current_line:
        lines.append(current_line)

    for i, line in enumerate(lines):
        cell = f'{start_col}{start_row + i}'
        ws[cell] = line
        ws[cell].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(f'{cell}:{chr(ord(start_col)+2)}{start_row + i}')
        ws[cell].font = Font(bold=False, size=12)
        number_of_rows_used += 1
        
    return number_of_rows_used

def get_sample_price(sample_type):
    global country, analysis
    
    if country == "Canada":
                
        price_dict_rush = {
            "A": 0.00,
            "BA": 100.00,
            "BAF": 120.00,
            "AOC": 100.00,
            "RCS": 0.00,
            "B": 65.00,
            "BF": 120.00,
            "S": 65.00,
            "SF": 120.00,
            "OT": 0.00
        }
        
        price_dict_standard = {
            "A": 0.00,
            "BA": 50.00,
            "BAF": 120.00,
            "AOC": 50.00,
            "RCS": 0.00,
            "B": 35.00,
            "BF": 120.00,
            "S": 35.00,
            "SF": 120.00,
            "OT": 0.00
        }
        
    else:
        
        price_dict_rush = {
            "A": 0.00,
            "BA": 75.00,
            "BAF": 90.00,
            "AOC": 75.00,
            "RCS": 0.00,
            "B": 50.00,
            "BF": 90.00,
            "S": 50.00,
            "SF": 90.00,
            "OT": 0.00
        }
        
        price_dict_standard = {
            "A": 0.00,
            "BA": 40.00,
            "BAF": 90.00,
            "AOC": 40.00,
            "RCS": 0.00,
            "B": 25.00,
            "BF": 90.00,
            "S": 25.00,
            "SF": 90.00,
            "OT": 0.00
        }
        
    if analysis == "Rush":
        return price_dict_rush.get(sample_type, 0.00)
    else:
        return price_dict_standard.get(sample_type, 0.00)

def get_sample_name(sample_type):
    global analysis
    
    matrix_code_dict = {
        "A": "Anderson Air",
        "BA": f"{analysis} Burkard Air",
        "BAF": f"{analysis} Burkard Air Fire-Related Particulate",
        "AOC": f"{analysis} Air-O-Cell",
        "RCS": "RCS Air",
        "B": f"{analysis} Bulk Growth & Spores",
        "BF": f"{analysis} Bulk Fire-Related Particulate",
        "S": f"{analysis} Surface Growth & Spores",
        "SF": f"{analysis} Surface Fire-Related Particulate",
        "OT": "Other Type"
    }
    return matrix_code_dict.get(sample_type, "Unknown")

def display_data():
    header()
    add_styling()
    
def header():
    global ws, country, starting_row, header_info
    row = starting_row
    
    ws[f'J{row}'] = "Invoice #" + invoice_number()
    ws[f'J{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'J{row}'].font = Font(bold=True, size=26)

    row = row + 2
    img = Image(get_asset_path('Assets/BSILogoUS.png') if country == 'United States' else get_asset_path('Assets/BSILogoCAN.png'))
    img.width = 275 #280
    img.height = 95 #101
    ws.merge_cells(f'A{row}:D{row+1}')  # Merges A-D cells, giving more room to center the image
    ws.add_image(img, f'A{row}')

    ws[f'E{row}'] = "In United States:"
    ws.merge_cells(f'E{row}:J{row}')
    ws[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'E{row}'].font = Font(bold=True, size=12)

    row = row + 1
    ws[f'E{row}'] = header_info['us_address']
    ws.merge_cells(f'E{row}:J{row}')
    ws[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'E{row}'].font = Font(bold=False, size=12)

    row = row + 2
    ws[f'E{row}'] = "In Canada:"
    ws.merge_cells(f'E{row}:J{row}')
    ws[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'E{row}'].font = Font(bold=True, size=12)
    
    row = row + 1
    ws[f'E{row}'] = header_info['can_address']
    ws.merge_cells(f'E{row}:J{row}')
    ws[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'E{row}'].font = Font(bold=False, size=12)
    
    row = row + 3
    line_across(row)
    
    row = row + 1
    client_job(row)

def client_job(row):
    global ws, client_job_info, handle_log_info
    ws[f'A{row}'] = "Attention"
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'A{row}'].font = Font(bold=True, size=12)

    ws[f'B{row}'] = client_job_info['contact']
    ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'B{row}:E{row}')
    ws[f'B{row}'].font = Font(bold=False, size=12)
    
    ws[f'F{row}'] = "Invoice Date"
    ws[f'F{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'F{row}'].font = Font(bold=True, size=12)

    ws[f'H{row}'] = handle_log_info['reviewed_by']['date_time']
    ws[f'H{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'H{row}:J{row}')
    ws[f'H{row}'].font = Font(bold=False, size=12)

    row = row + 1
    ws[f'A{row}'] = "Address"
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'A{row}'].font = Font(bold=True, size=12)

    ws[f'B{row}'] = client_job_info['client']
    ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'B{row}:E{row}')
    ws[f'B{row}'].font = Font(bold=False, size=12)

    ws[f'F{row}'] = "Invoice No."
    ws[f'F{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'F{row}'].font = Font(bold=True, size=12)

    ws[f'H{row}'] = invoice_number()
    ws[f'H{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'H{row}:I{row}')
    ws[f'H{row}'].font = Font(bold=False, size=12)

    row = row + 1
    starting_address_row = row
    ws[f'F{row}'] = "Project No."
    ws[f'F{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'F{row}'].font = Font(bold=True, size=12)

    ws[f'H{row}'] = project_number()
    ws[f'H{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'H{row}:I{row}')
    ws[f'H{row}'].font = Font(bold=False, size=12)

    row = row + 1
    ws[f'F{row}'] = "Client No."
    ws[f'F{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws[f'F{row}'].font = Font(bold=True, size=12)

    ws[f'H{row}'] = str(client_job_info['job_#'])
    ws[f'H{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'H{row}:I{row}')
    ws[f'H{row}'].font = Font(bold=False, size=12)
    
    
    address_rows = split_text_based_on_column_width(f'B{starting_address_row}', client_job_info['address'], 25)
    row = starting_address_row + (2 if address_rows == 1 else address_rows)
    
    row = row + 1
    line_across(row)
    
    row = row + 1
    content(row)

def content(row):
    global ws, client_job_info, sample_info
    ws[f'A{row}'] = content_title()['first']
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{row}'].font = Font(bold=True, size=12)

    row = row + 1
    ws[f'A{row}'] = content_title()['second']
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{row}'].font = Font(bold=True, size=12)
    
    row = row + 3
    ws[f'A{row}'] = 'Quantity'
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{row}'].font = Font(bold=True, size=12)
    
    ws[f'B{row}'] = 'Sample Type'
    ws.merge_cells(f'B{row}:D{row}')
    ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'B{row}'].font = Font(bold=True, size=12)

    ws[f'F{row}'] = 'Rate'
    ws[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'F{row}'].font = Font(bold=True, size=12)
    
    ws[f'I{row}'] = 'Billed Amount'
    ws[f'I{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws[f'I{row}'].font = Font(bold=True, size=12)
    
    for col in range(1,10):
        ws.cell(row=row, column=col).border = Border(bottom=Side(style='thin'))
        
    row = row + 2
    generate_sample_rows(row)
    
def generate_sample_rows(row):
    global sample_info
    start_row = row
    total_billed = 0

    sample_summary = {}

    for sample in sample_info:
        sample_type = sample['matrix']

        if sample_type in sample_summary:
            sample_summary[sample_type] += 1
        else:
            sample_summary[sample_type] = 1

    for sample_type, quantity in sample_summary.items():
        sample_name = get_sample_name(sample_type)
        rate = get_sample_price(sample_type)

        total_price = quantity * rate
        total_billed += total_price  

        # Insert values into the spreadsheet for each distinct sample type
        ws[f'A{row}'] = quantity
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws[f'A{row}'].font = Font(size=12)
        
        ws[f'B{row}'] = sample_name
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'].alignment = Alignment(horizontal='left')
        ws[f'B{row}'].font = Font(size=12)
        
        ws[f'F{row}'] = rate  # Insert as a number
        ws[f'F{row}'].alignment = Alignment(horizontal='center')
        ws[f'F{row}'].font = Font(size=12)
        ws[f'F{row}'].number_format = '$ #,##0.00'  # Currency format
        
        ws[f'I{row}'] = f"=A{row} * F{row}"  # Insert as a number
        ws[f'I{row}'].alignment = Alignment(horizontal='center')
        ws[f'I{row}'].font = Font(size=12)
        ws[f'I{row}'].number_format = '$ #,##0.00'  # Currency format

        row = row + 1 
        
    end_row = row - 1
    
    # Billed Amount Sum
    ws[f'I{row}'] = f"=SUM(I{start_row}:I{end_row})"
    ws[f'I{row}'].alignment = Alignment(horizontal='center')
    ws[f'I{row}'].font = Font(size=12)
    ws[f'I{row}'].number_format = '$ #,##0.00'  # Currency format
    ws.cell(row=row, column=9).border = Border(top=Side(style='double'))
    row = row + 1

    # HST
    total_billed_row = row - 1
    if (country == "Canada"):
        ws[f'F{row}'] = "H.S.T."
        ws[f'F{row}'].alignment = Alignment(horizontal='right')
        ws[f'F{row}'].font = Font(size=12)

        ws[f'I{row}'] = f"=I{total_billed_row} * {0.13}"  # Insert as a number
        ws[f'I{row}'].alignment = Alignment(horizontal='center')
        ws[f'I{row}'].font = Font(size=12)
        ws[f'I{row}'].number_format = '$ #,##0.00'  # Currency format
        row = row + 1
    
    # Final Total Activity
    row = row + 1
    ws.cell(row=row, column=5).border = Border(left=Side(style='double'), top=Side(style='double'), bottom=Side(style='double'))
    ws.cell(row=row, column=9).border = Border(right=Side(style='double'), top=Side(style='double'), bottom=Side(style='double'))
    for col in range(6,9):
        ws.cell(row=row, column=col).border = Border(top=Side(style='double'), bottom=Side(style='double'))
        
    ws[f'E{row}'] = "Final Total Activity"
    ws.merge_cells(f'E{row}:G{row}')
    ws[f'E{row}'].alignment = Alignment(horizontal='right')
    ws[f'E{row}'].font = Font(bold=True, size=12)
    
    hst_row = row - 1
    ws[f'I{row}'] = f"=SUM(I{total_billed_row}: I{hst_row})"  # Insert as a number
    ws[f'I{row}'].alignment = Alignment(horizontal='center')
    ws[f'I{row}'].font = Font(bold=True, size=12)
    ws[f'I{row}'].number_format = '$ #,##0.00'  # Currency format
    
    row = row + 10
    footer(row)

def footer(row):
    global ws, end_row
    
    ws[f'A{row}'] = "If you have any questions regarding this invoice, please do not hesitate"
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{row}'].font = Font(bold=True, size=12)

    row = row + 1
    ws[f'A{row}'] = "to contact Mr. Jonathan Solomon."
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{row}'].font = Font(bold=True, size=12)

    row = row + 2
    ws[f'A{row}'] = "Please return a copy of this invoice with remittance."
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{row}'].font = Font(bold=True, size=12)

    row = row + 2
    ws[f'A{row}'] = "Terms: Payment on Receipt of Invoice - Interest of 1½% per month compounded monthly"
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{row}'].font = Font(bold=True, size=12)

    row = row + 1
    ws[f'A{row}'] = "(19.56% annually) payable on amounts unpaid within 30 days."
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{row}'].font = Font(bold=True, size=12)
    
    row = row + 3
    end_row = row
    line_across(row)

def generate_invoice(data, global_country, global_type):
    
    global header_info, client_job_info, sample_info, handle_log_info
    header_info = data['header_info']
    client_job_info = data['client_job_info']
    sample_info = data['sample_info']
    handle_log_info = data['handle_log_info']
    
    global ws, country, type, analysis
    analysis = str(client_job_info['analysis'])
    country = global_country
    type = global_type
    
    global starting_row
    starting_row = 1 # Starting row of Invoice
    
    wb = Workbook()
    ws = wb.active
    
    setup_excel()
    display_data()
    
    if platform.system() == 'Windows':
        desktop_path = f'{os.path.expanduser("~")}/OneDrive/Desktop/'
    elif platform.system() == 'Darwin':
        desktop_path = f'{os.path.expanduser("~")}/Desktop/'
        
    job_num = client_job_info['job_#']
    project = get_before_comma(client_job_info['project'])
    invoice_num = invoice_number()
    folder_name = f'{job_num} - {project}'
    folder_path = f'{desktop_path}{folder_name}'
    
    os.makedirs(folder_path, exist_ok=True)
    output_path = f'{folder_path}/{invoice_num} Invoice {project}.xlsx'
    wb.save(output_path)
    print(f"Invoice saved to {output_path}")